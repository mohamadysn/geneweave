#!/usr/bin/python3
#-*- coding: utf-8 -*-

import csv
import json
import re

COLUMN_HEADERS = [
    "Gene Symbol and Organism",
    "NCBI Official Full Name",
    "NCBI Gene ID",
    "NCBI RefSeq RNA ID",
    "NCBI RefSeq Protein ID",
    "Ensembl Gene ID",
    "Ensembl Genome Browser",
    "Ensembl Transcript ID",
    "Ensembl Protein ID",
    "Ensembl Orthologs",
    "UniProt ID",
    "UniProt Official Protein Name",
    "STRING ID",
    "STRING Network",
    "ProSite Motifs",
    "ProSite Graphical View",
    "Pfam Domains",
    "Pfam Graphical View",
    "PDB Structure",
    "KEGG ID",
    "KEGG Pathway",
    "GO Molecular Function",
    "GO Cellular Component",
    "GO Biological Process",
]


def cell_to_text(cell):
    if cell in ("No data found", ["No data found"], None, []):
        return "No data found"

    text = str(cell)
    text = re.sub(r"<br\s*/?>", " | ", text, flags=re.IGNORECASE)
    text = re.sub(r'<a\s+href="[^"]*"[^>]*>([^<]*)</a>', r"\1", text, flags=re.IGNORECASE)
    text = re.sub(r"[\[\]']", "", text)
    return re.sub(r"\s+", " ", text).strip()


def rows_as_text(infos):
    return [[cell_to_text(cell) for cell in row] for row in infos]


def write_csv(infos, output_path):
    with open(output_path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow(COLUMN_HEADERS)
        writer.writerows(rows_as_text(infos))


def write_excel(infos, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as error:
        raise ImportError("Install openpyxl: pip install openpyxl") from error

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Annotations"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)

    sheet.append(COLUMN_HEADERS)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows_as_text(infos):
        sheet.append(row)

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    sheet.freeze_panes = "A2"
    workbook.save(output_path)


def write_json(infos, output_path, summary=None, metadata=None):
    payload = {
        "columns": COLUMN_HEADERS,
        "rows": rows_as_text(infos),
    }
    if summary:
        payload["summary"] = summary
    if metadata:
        payload["metadata"] = metadata
    with open(output_path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def write_manifest(output_path, paths, summary, metadata):
    manifest = {
        "version": metadata.get("version"),
        "generated_at": metadata.get("generated_at"),
        "duration_seconds": summary.get("duration_seconds"),
        "total_genes": summary.get("total_genes"),
        "coverage_percent": summary.get("coverage_percent"),
        "errors": summary.get("errors", []),
        "outputs": {
            key: paths[key]
            for key in ("html", "csv", "xlsx", "json", "summary")
            if key in paths
        },
    }
    with open(output_path, "w", encoding="utf-8") as handle:
        json.dump(manifest, handle, ensure_ascii=False, indent=2)

